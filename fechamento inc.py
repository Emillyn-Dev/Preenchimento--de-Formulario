# =============================================================
# Como usar:
#   1. Abra o terminal na pasta do script
#   2. Ative o venv:  .\.venv\Scripts\Activate
#   3. Execute:       python f_inc_2.py
#
#   Parâmetros opcionais (sem precisar editar o código):
#     python f_inc_2.py --planilha "RITM - INC.xlsx" --sheet INC
# =============================================================

from __future__ import annotations

import argparse
import logging
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait


# PASTA_RAIZ = pasta onde este arquivo .py esta salvo. Como a planilha e o
# msedgedriver.exe ficam na mesma pasta do codigo, os caminhos abaixo se
# ajustam sozinhos, nao importa em qual maquina/pasta o projeto esteja.
PASTA_RAIZ = Path(__file__).resolve().parent

PLANILHA_PADRAO = PASTA_RAIZ / "RITM - INC.xlsx"  # [MUDAR AQUI, se o nome do arquivo for diferente]
SHEET_PADRAO = "INC"  # troque pra "RITM" se quiser fechar os chamados dessa aba
DRIVER_PADRAO = PASTA_RAIZ / "edgedriver_win64" / "msedgedriver.exe"  # [MUDAR AQUI, se o nome do arquivo for diferente]
# A versao do msedgedriver.exe tem que bater com a versao do Edge
# instalado na maquina. Se for diferente, baixe o driver certo em:
# https://developer.microsoft.com/microsoft-edge/tools/webdriver/

# O ServiceNow dessa empresa usa autenticacao integrada do Windows (SSO):
# ele reconhece sozinho o usuario que esta logado no Windows da maquina,
# sem precisar de cookie/login salvo no perfil do Edge. Por isso o script
# so abre uma sessao de automacao comum do Edge, que ja entra logada
# sozinha, em qualquer conta que estiver ativa na maquina -- assim funciona
# igual em qualquer computador, sem precisar fechar o Edge antes de rodar.


URL_LISTA = (
    "https://rdslprod.service-now.com/now/nav/ui/classic/params/target/"
    "task_list.do%3Fsysparm_view%3D%26sysparm_query%3Dassignment_group%253Dc87ddb681b907450bcd687bfe54bcb4f"
    "%255EORassignment_group%253Dd07ddb681b907450bcd687bfe54bcb8b"
    "%255EORassignment_group%253D3a69b0171b6055904e95a68fe54bcbb3"
    "%255EORassignment_group%253D7f6d9b681b907450bcd687bfe54bcb48"
    "%255EORassignment_group%253De87d5f681b907450bcd687bfe54bcb2b"
    "%255EORassignment_group%253Dd07d1f681b907450bcd687bfe54bcb02"
    "%255EstateNOT%2520IN3%252C4%252C7%252C8%252C9%252C6%252C106%252C107%252C157"
    "%255Eassignment_group%253Dc87ddb681b907450bcd687bfe54bcb4f"
    "%26sysparm_fixed_query%3D"
)


ID_BUSCA_GLOBAL = "sncwsgs-typeahead-input"


ID_CAMPO_GRUPO = "sys_display.incident.assignment_group"
ID_CAMPO_FECHAR = "sys_display.incident.assigned_to"

# Campo "Estado" do chamado (select). Se ja estiver como "Resolvido", o
# chamado e pulado sem mexer em nada -- so grava na planilha que ja estava
# fechado e segue pro proximo.
ID_CAMPO_ESTADO = "incident.state"
TEXTO_ESTADO_RESOLVIDO = "resolvido"

# Nome do template que sempre deve ser aplicado no final (fixo, nao vem da planilha).
# O "data-ref" e o identificador interno (sys_id) do template -- casar por ele e
# mais confiavel do que casar pelo texto, porque nao muda mesmo que o nome do
# template seja editado depois.
NOME_TEMPLATE = "Emy - INC"
REF_TEMPLATE = "1995947d970a435848cfbcc6f053aff8"

# Id do iframe onde o ServiceNow (UI classica) carrega o formulario do
# chamado (campos, barra de templates, botao Resolver). Sem trocar pra esse
# iframe, o Selenium fica "olhando" pro documento principal e nao acha
# nenhum desses elementos.
ID_IFRAME_FORMULARIO = "gsft_main"

# Nome da coluna onde o resultado de cada chamado e gravado de volta na

COLUNA_STATUS = "status"

# =========================
# LOGGING
# =========================
# Substitui os "print" espalhados por um logger, so no terminal mesmo
# (mais organizado, com hora e nivel de cada mensagem).
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("f_inc")


# =========================
# MODELO DE DADOS
# =========================
@dataclass
class Chamado:
    numero: str
    linha_excel: int  # numero da linha na planilha (1-indexado, igual ao Excel) -- usado pra gravar o status de volta
    grupo: Optional[str] = None
    fechar: Optional[str] = None


# =========================
# LEITURA DA PLANILHA
# =========================
def normalizar(texto) -> str:
    return (
        unicodedata.normalize("NFKD", str(texto))
        .encode("ascii", errors="ignore")
        .decode("utf-8")
        .lower()
        .strip()
    )


def valor_valido(val) -> bool:
    return pd.notna(val) and str(val).strip().lower() not in ("", "nan", "none")


def verificar_planilha_fechada(planilha: Path) -> None:
    """Falha rapido e com mensagem clara se o Excel estiver com o arquivo aberto."""
    try:
        with open(planilha, "r+b"):
            pass
    except FileNotFoundError:
        log.error("Planilha nao encontrada: %s", planilha)
        raise SystemExit(1)
    except PermissionError:
        log.warning("Feche o arquivo no Excel antes de rodar o script!")
        input("Pressione ENTER depois de fechar o arquivo...")


def carregar_chamados(planilha: Path, sheet: str) -> tuple[list[Chamado], int]:
    """Le a planilha, detecta a linha de cabecalho automaticamente e devolve
    a lista de chamados validos a processar, junto com o indice (0-based)
    da linha de cabecalho -- usado depois pra gravar o status na linha
    certa do Excel."""
    df_raw = pd.read_excel(planilha, sheet_name=sheet, header=None, dtype=str)

    header_row = None
    for i, row in df_raw.iterrows():
        colunas = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip() != ""]
        if len(colunas) >= 3:
            header_row = i
            break

    if header_row is None:
        log.error("Nao foi possivel detectar o cabecalho da planilha.")
        raise SystemExit(1)

    df = pd.read_excel(planilha, sheet_name=sheet, header=header_row, dtype=str)
    df.columns = [normalizar(col) for col in df.columns]
    # NAO usar reset_index(drop=True) aqui: o indice do pandas precisa
    # continuar batendo com a posicao real na planilha, pra calcular
    # linha_excel corretamente logo abaixo.
    df = df.dropna(how="all")

    if "chamado" not in df.columns:
        log.error("A planilha nao tem a coluna 'chamado'. Nada a fazer.")
        raise SystemExit(1)

    # linha_excel: header_row e 0-based e aponta pra linha do cabecalho;
    # os dados comecam na linha seguinte. Somando +2 (1 pra virar 1-based,
    # +1 pra pular o cabecalho) e o indice da linha no df, chegamos na
    # linha exata do Excel (1-indexada) pra cada chamado.
    #
    # So entra na lista quem tem a coluna 'chamado' preenchida E a coluna
    # 'status' (nome definido em COLUNA_STATUS) ainda vazia na mesma linha
    # -- ou seja, chamados ja processados numa rodada anterior sao pulados
    # automaticamente.
    coluna_status_normalizada = normalizar(COLUNA_STATUS)
    chamados = [
        Chamado(
            numero=str(linha["chamado"]).strip(),
            linha_excel=header_row + 2 + idx,
            grupo=str(linha.get("grupo", "")).strip() if valor_valido(linha.get("grupo", "")) else None,
            fechar=str(linha.get("fechar", "")).strip() if valor_valido(linha.get("fechar", "")) else None,
        )
        for idx, linha in df.iterrows()
        if valor_valido(linha.get("chamado", ""))
        and str(linha["chamado"]).strip().upper() != "N/A"
        and not valor_valido(linha.get(coluna_status_normalizada, ""))
    ]

    pulados = int(df["chamado"].apply(valor_valido).sum()) - len(chamados)
    if pulados > 0:
        log.info("%d chamado(s) ja tinham status preenchido e foram pulados.", pulados)

    return chamados, header_row


# =========================
# AUTOMACAO SELENIUM
# =========================
class AutomacaoServiceNow:
    """Agrupa o driver e as esperas junto com todas as acoes na tela, pra
    nao depender de variaveis globais espalhadas pelo modulo."""

    def __init__(self, driver: webdriver.Edge):
        self.driver = driver
        self.wait_longo = WebDriverWait(driver, 60)
        self.wait_medio = WebDriverWait(driver, 15)
        self.wait_curto = WebDriverWait(driver, 5)

    # ---------- utilidades ----------
    JS_BUSCA_PROFUNDA = """
        function buscaProfunda(raiz, seletor) {
            let el = raiz.querySelector(seletor);
            if (el) return el;
            const todos = raiz.querySelectorAll('*');
            for (const no of todos) {
                if (no.shadowRoot) {
                    el = buscaProfunda(no.shadowRoot, seletor);
                    if (el) return el;
                }
            }
            return null;
        }
        return buscaProfunda(document, arguments[0]);
    """

    def buscar_elemento_profundo(self, seletor_css: str) -> Optional[WebElement]:
        """Procura um elemento pelo seletor CSS em todo o documento,
        incluindo dentro de Shadow DOM aberto. A UI 'Next Experience' do
        ServiceNow encapsula componentes (como a busca global) dentro de
        shadow roots, e o find_element(By.ID, ...) padrao do Selenium nao
        consegue enxergar la dentro -- por isso essa busca alternativa."""
        return self.driver.execute_script(self.JS_BUSCA_PROFUNDA, seletor_css)

    def clicar(self, elemento: WebElement) -> None:
        try:
            elemento.click()
        except Exception:
            self.driver.execute_script(
                "arguments[0].scrollIntoView(true); arguments[0].click();", elemento
            )

    def aguardar_pagina_completa(self) -> None:
        """Espera o carregamento completo da pagina (document.readyState == complete)."""
        log.info("Aguardando pagina carregar completamente...")
        self.wait_longo.until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(0.1)  # pequena folga pra a tela terminar de montar depois do "complete"

    def entrar_frame_formulario(self) -> bool:
        """Troca o contexto do Selenium pra dentro do iframe onde o
        ServiceNow classico carrega o formulario do chamado (campos, barra
        de templates, botao Resolver). Sem isso, os find_element/find_by_id
        continuam procurando no documento principal e nunca acham nada --
        falham calados, so com um aviso no log.

        Na interface "Next Experience" do ServiceNow, esse iframe
        ('gsft_main') fica dentro de um Web Component com Shadow DOM
        (slot="core-ui"). O metodo padrao do Selenium
        (frame_to_be_available_and_switch_to_it) so enxerga o documento
        principal e nunca acha esse iframe -- por isso primeiro tentamos o
        jeito padrao (mais rapido, cobre o caso da UI classica "pura") e,
        se falhar, caimos pra busca profunda em Shadow DOM (igual a usada
        no campo de busca global)."""
        self.driver.switch_to.default_content()

        # Tentativa 1: jeito padrao (funciona se o iframe estiver no documento principal, sem Shadow DOM)
        try:
            self.wait_curto.until(
                EC.frame_to_be_available_and_switch_to_it((By.ID, ID_IFRAME_FORMULARIO))
            )
            log.info("Entrou no iframe do formulario ('%s').", ID_IFRAME_FORMULARIO)
            return True
        except Exception:
            self.driver.switch_to.default_content()

        # Tentativa 2: busca profunda em Shadow DOM (Next Experience)
        log.info("Iframe nao encontrado no documento principal -- procurando dentro de Shadow DOM...")
        seletor = f"#{ID_IFRAME_FORMULARIO}"
        prazo = time.time() + 15
        elemento_iframe = None
        while time.time() < prazo:
            elemento_iframe = self.buscar_elemento_profundo(seletor)
            if elemento_iframe is not None:
                break
            time.sleep(0.5)

        if elemento_iframe is None:
            log.warning(
                "Nao consegui entrar no iframe do formulario ('%s'): nao encontrado nem no documento "
                "principal nem em Shadow DOM.",
                ID_IFRAME_FORMULARIO,
            )
            return False

        try:
            self.driver.switch_to.frame(elemento_iframe)
            log.info("Entrou no iframe do formulario ('%s') via Shadow DOM.", ID_IFRAME_FORMULARIO)
            return True
        except Exception as e:
            log.warning("Achei o iframe em Shadow DOM mas nao consegui trocar de contexto: %s", e)
            return False

    def sair_frame_formulario(self) -> None:
        """Volta o contexto do Selenium pro documento principal (fora de
        qualquer iframe) -- necessario antes de usar a busca global de novo."""
        self.driver.switch_to.default_content()

    def selecionar_na_lista(self, numero: str) -> bool:
        """Se o Enter levar pra uma pagina de lista de resultados (em vez de
        abrir o chamado direto), procura um link cujo texto contenha o
        numero exato do chamado e clica nele."""
        try:
            links = self.driver.find_elements(By.PARTIAL_LINK_TEXT, numero)
            visiveis = [l for l in links if l.is_displayed()]
            if visiveis:
                log.info("Lista de resultados encontrada -- clicando em '%s'...", numero)
                self.clicar(visiveis[0])
                return True
        except Exception:
            pass
        return False

    def selecionar_primeira_sugestao(self, campo: WebElement) -> bool:
        """Seleciona a primeira sugestao da lista de autocomplete classica do
        ServiceNow. Primeiro tenta achar a lista pelo id indicado no
        atributo 'aria-owns' do proprio campo; se nao achar nada visivel,
        usa teclado (seta pra baixo + Enter) como alternativa."""
        aria_owns = campo.get_attribute("aria-owns")
        if aria_owns:
            try:
                opcoes = self.wait_curto.until(
                    lambda d: [
                        o
                        for o in d.find_elements(By.CSS_SELECTOR, f"#{aria_owns} li, #{aria_owns} a")
                        if o.is_displayed() and o.text.strip()
                    ]
                    or False
                )
                self.clicar(opcoes[0])
                log.info("Selecionado (lista): '%s'", opcoes[0].text.strip())
                return True
            except Exception:
                pass

        # Fallback: seta pra baixo + Enter (destaca e confirma a 1a sugestao)
        try:
            campo.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.2)
            campo.send_keys(Keys.ENTER)
            log.info("Selecionado via teclado (seta + Enter).")
            return True
        except Exception as e:
            log.warning("Nao consegui selecionar nenhuma sugestao: %s", e)
            return False

    def preencher_autocomplete_classic(self, id_campo: str, texto: str) -> bool:
        """Preenche um campo de referencia classico do ServiceNow (tipo
        'assignment_group' ou 'assigned_to'), espera as sugestoes
        carregarem via AJAX e seleciona a primeira."""
        log.info("Preenchendo campo '%s' com '%s'...", id_campo, texto)
        try:
            campo = self.wait_medio.until(EC.element_to_be_clickable((By.ID, id_campo)))
        except Exception as e:
            log.warning("Campo '%s' nao encontrado: %s", id_campo, e)
            return False

        self.clicar(campo)
        campo.send_keys(Keys.CONTROL, "a")
        campo.send_keys(Keys.DELETE)
        time.sleep(0.15)
        campo.send_keys(texto)
        time.sleep(0.5)  # espera o AJAX trazer as sugestoes

        return self.selecionar_primeira_sugestao(campo)

    def aplicar_template(self, nome_template: str = NOME_TEMPLATE, ref_template: Optional[str] = REF_TEMPLATE) -> bool:
        """Procura, entre os templates disponiveis na tela, o link certo e
        clica nele. Da preferencia a casar pelo 'data-ref' (o sys_id do
        template, que nao muda mesmo se o nome for editado depois); se nao
        vier ref ou nao achar por ele, cai pra casar pelo texto do nome."""
        nome_lower = nome_template.strip().lower()

        try:
            links = self.wait_medio.until(
                lambda d: [
                    l
                    for l in d.find_elements(By.CSS_SELECTOR, "a.template-item[data-type='apply-template']")
                    if l.is_displayed()
                ]
                or False
            )
        except Exception:
            log.warning("Nenhum template apareceu na tela pra aplicar.")
            return False

        if ref_template:
            for link in links:
                if (link.get_attribute("data-ref") or "").strip().lower() == ref_template.strip().lower():
                    log.info("Aplicando template (por ref): '%s'", link.text.strip())
                    self.clicar(link)
                    self._aguardar_template_aplicado(link)
                    return True

        for link in links:
            if nome_lower in link.text.strip().lower():
                log.info("Aplicando template (por nome): '%s'", link.text.strip())
                self.clicar(link)
                self._aguardar_template_aplicado(link)
                return True

        log.warning("Nao achei um template com '%s' entre os disponiveis:", nome_template)
        for link in links:
            log.warning("  - %s (ref: %s)", link.text.strip(), link.get_attribute("data-ref"))
        return False

    def _aguardar_template_aplicado(self, link_clicado: WebElement, tempo_limite: float = 15.0) -> None:
        """Espera o template realmente terminar de ser aplicado antes de
        seguir. O ServiceNow processa isso via AJAX (preenche os campos e
        fecha o painel de templates), e o sinal mais confiavel que temos
        e o proprio link clicado sumir do DOM ('ficar stale') quando o
        painel fecha. Alem disso, colocamos uma folga extra depois, porque
        o preenchimento dos campos pode continuar por um instante mesmo
        depois do painel fechar."""
        log.info("Aguardando o template ser aplicado...")
        try:
            WebDriverWait(self.driver, tempo_limite).until(EC.staleness_of(link_clicado))
            log.info("Template aplicado (painel de templates fechou).")
        except Exception:
            log.info("Nao confirmei o fechamento do painel de templates -- seguindo apos pausa de seguranca.")
        time.sleep(0.1)  # folga extra pro AJAX terminar de preencher os campos

    def aguardar_formulario_pronto(self, tempo_limite: float = 20.0) -> None:
        """Espera o formulario classico terminar de inicializar o proprio
        JavaScript (g_form.isLoaded()) antes de interagir com ele. So a
        pagina estar com readyState 'complete' nao garante que os botoes
        (como 'Resolver') ja tem o evento de clique conectado -- clicar
        cedo demais faz o clique "sumir" sem erro nenhum, sem popup, sem
        nada mudar na tela."""
        try:
            WebDriverWait(self.driver, tempo_limite).until(
                lambda d: d.execute_script(
                    "try { return typeof g_form !== 'undefined' && g_form.isLoaded && g_form.isLoaded(); } "
                    "catch (e) { return false; }"
                )
            )
            log.info("Formulario (g_form) pronto.")
        except Exception:
            log.info("Nao confirmei que g_form terminou de carregar -- seguindo mesmo assim apos pausa extra.")
            time.sleep(0.15)

    def obter_texto_estado(self) -> Optional[str]:
        """Le o valor atualmente selecionado no campo 'Estado'
        (select#incident.state) do formulario do chamado. Precisa ser
        chamado depois de ja estar dentro do iframe do formulario. Devolve
        o texto normalizado (sem acento, minusculo) ou None se nao
        conseguir ler o campo."""
        try:
            elemento = self.wait_curto.until(
                EC.presence_of_element_located((By.ID, ID_CAMPO_ESTADO))
            )
            texto = Select(elemento).first_selected_option.text
            log.info("Campo 'Estado' atual: '%s'", texto.strip())
            return normalizar(texto)
        except Exception as e:
            log.info("Nao consegui ler o campo 'Estado' (%s) -- seguindo o fluxo normal.", e)
            return None

    def clicar_resolver(self) -> bool:
        """Clica no botao "Resolver" do formulario do chamado (id:
        resolve_incident) e confirma que a URL mudou."""
        try:
            botao = self.wait_medio.until(EC.element_to_be_clickable((By.ID, "resolve_incident")))
        except Exception as e:
            log.warning("Botao 'Resolver' nao encontrado: %s", e)
            return False

        # (O formulario ja foi confirmado como pronto -- g_form.isLoaded() -- antes
        # de comecar a preencher os campos, entao nao precisa checar de novo aqui.)

        url_antes = self.driver.current_url

        # Clique reforcado: tenta o clique normal/JS (via self.clicar) e,
        # se a URL nao mudar em alguns segundos, tenta de novo com
        # ActionChains (simula um clique de mouse "de verdade", que
        # dispara os eventos que alguns handlers do ServiceNow esperam).
        self.clicar(botao)
        log.info("Botao 'Resolver' clicado.")

        try:
            WebDriverWait(self.driver, 5).until(lambda d: d.current_url != url_antes)
        except Exception:
            log.info("Nada mudou apos o clique normal -- tentando de novo com ActionChains...")
            try:
                botao = self.wait_curto.until(EC.element_to_be_clickable((By.ID, "resolve_incident")))
                ActionChains(self.driver).move_to_element(botao).pause(0.2).click().perform()
                log.info("Botao 'Resolver' clicado via ActionChains.")
            except Exception as e:
                log.warning("Nao consegui clicar em 'Resolver' via ActionChains: %s", e)

        try:
            self.wait_medio.until(lambda d: d.current_url != url_antes)
            log.info("Chamado resolvido -- voltou pra pagina de tarefas.")
            # Ao resolver, o iframe do formulario e destruido e a pagina de
            # tarefas (antiga) carrega no documento principal. Saimos do
            # contexto do iframe (senao ele fica "stale") e esperamos essa
            # pagina terminar de carregar antes de seguir pro proximo chamado.
            self.driver.switch_to.default_content()
            self.aguardar_pagina_completa()
            return True
        except Exception:
            log.warning("Nao confirmei se voltou pra pagina de tarefas. Verifique manualmente.")
            self.driver.switch_to.default_content()
            return False

    def abrir_busca_global(self) -> WebElement:
        """O campo de busca global so fica realmente "ativo" depois que voce
        clica nele. Como e um componente React/Web Component (as vezes
        dentro de Shadow DOM), um clique via JS as vezes nao da foco de
        teclado de verdade, entao confirmamos o foco explicitamente e
        forcamos via JS se precisar."""
        seletor = f"#{ID_BUSCA_GLOBAL}"
        campo = None
        prazo = time.time() + 60
        while time.time() < prazo:
            campo = self.buscar_elemento_profundo(seletor)
            if campo is not None:
                break
            time.sleep(0.5)

        if campo is None:
            raise TimeoutError(
                f"Campo de busca global ({seletor}) nao encontrado em 60s "
                "(procurado inclusive dentro de Shadow DOM)."
            )

        self.clicar(campo)
        time.sleep(0.3)

        tem_foco = self.driver.execute_script(
            "return (arguments[0].getRootNode().activeElement || document.activeElement) === arguments[0];",
            campo,
        )
        if not tem_foco:
            log.info("O clique nao deu foco real no campo -- forcando foco via JS.")
            self.driver.execute_script("arguments[0].focus();", campo)
            time.sleep(0.3)
            tem_foco = self.driver.execute_script(
                "return (arguments[0].getRootNode().activeElement || document.activeElement) === arguments[0];",
                campo,
            )
            log.info("Foco apos forcar via JS: %s", tem_foco)

        try:
            self.wait_curto.until(
                lambda d: self.buscar_elemento_profundo(seletor).get_attribute("aria-expanded") == "true"
            )
            log.info("Dropdown de busca aberto (campo ativo).")
        except Exception:
            log.info("Nao confirmei se o dropdown de sugestoes abriu. Vou tentar digitar mesmo assim.")

        return self.buscar_elemento_profundo(seletor)

    def _digitar_numero(self, campo: WebElement, numero: str) -> str:
        """Digita o numero do chamado no campo de busca, com um fallback
        via ActionChains (digitacao caractere a caractere) para o caso do
        campo React ignorar o send_keys em lote."""
        campo.send_keys(Keys.CONTROL, "a")
        campo.send_keys(Keys.DELETE)
        time.sleep(0.15)
        campo.send_keys(numero)
        time.sleep(0.8)  # da tempo do typeahead reagir

        valor_atual = campo.get_attribute("value")
        if valor_atual and numero in valor_atual:
            return valor_atual

        log.info("Campo veio vazio -- tentando de novo com foco forcado + digitacao lenta...")
        self.driver.execute_script("arguments[0].focus();", campo)
        time.sleep(0.3)
        actions = ActionChains(self.driver)
        actions.click(campo)
        for caractere in numero:
            actions.send_keys(caractere)
            actions.pause(0.08)
        actions.perform()
        time.sleep(1.5)
        return campo.get_attribute("value")

    def abrir_chamado(self, numero: str) -> bool:
        """Vai para a pagina de tarefas, cola o numero do chamado na busca
        global, da Enter e confirma se abriu (direto ou por uma lista de
        resultados)."""
        log.info("Abrindo chamado %s...", numero)

        self.sair_frame_formulario()  # garante que comeca fora de qualquer iframe do chamado anterior
        self.driver.get(URL_LISTA)
        self.aguardar_pagina_completa()

        try:
            campo = self.abrir_busca_global()
        except Exception as e:
            log.warning("Campo de busca global nao encontrado: %s", e)
            return False

        time.sleep(0.5)  # pequena folga antes de digitar o numero (o campo ja foi confirmado como ativo)
        valor_atual = self._digitar_numero(campo, numero)
        log.debug("Valor no campo apos digitar: '%s'", valor_atual)

        log.info("Selecionando a primeira sugestao da lista...")
        campo.send_keys(Keys.ARROW_DOWN)
        time.sleep(0.2)
        campo.send_keys(Keys.ENTER)

        log.info("Aguardando resposta da busca...")
        return self._aguardar_abertura_chamado(numero)

    def _aguardar_abertura_chamado(self, numero: str, tempo_limite: float = 15.0) -> bool:
        """Confirma se o chamado abriu, tentando varios sinais em loop:

        1. A URL passou a conter o numero do chamado (as vezes acontece,
           mas o ServiceNow costuma navegar usando o sys_id interno em vez
           do numero, entao esse sinal sozinho nao e confiavel).
        2. O formulario do chamado carregou dentro do iframe (sinal mais
           forte -- se o campo de grupo esta la, o chamado abriu direto).
        3. Apareceu uma lista de resultados em vez de abrir direto -- nesse
           caso clica no link certo da lista.
        """
        prazo = time.time() + tempo_limite
        while time.time() < prazo:
            if numero.lower() in self.driver.current_url.lower():
                log.info("Foi direto para o chamado %s (confirmado pela URL).", numero)
                return True

            if numero.lower() in (self.driver.title or "").lower():
                log.info("Foi direto para o chamado %s (confirmado pelo titulo da aba).", numero)
                return True

            try:
                self.driver.switch_to.default_content()
                if self.driver.find_elements(By.ID, ID_IFRAME_FORMULARIO):
                    self.driver.switch_to.frame(ID_IFRAME_FORMULARIO)
                    formulario_presente = bool(self.driver.find_elements(By.ID, ID_CAMPO_GRUPO))
                    self.driver.switch_to.default_content()
                    if formulario_presente:
                        log.info("Foi direto para o chamado %s (confirmado pelo formulario).", numero)
                        return True
            except Exception:
                self.driver.switch_to.default_content()

            if self.selecionar_na_lista(numero):
                time.sleep(2)
                log.info("Chamado %s aberto a partir da lista.", numero)
                return True

            time.sleep(0.5)

        log.warning("Nao consegui confirmar se o chamado %s foi aberto. Verifique manualmente.", numero)
        return False

    def processar_chamado(self, item: Chamado) -> str:
        """Executa o fluxo completo (abrir, preencher, aplicar template,
        resolver) para um unico chamado.

        Devolve um status em texto:
          - "fechado"     -> o script resolveu o chamado agora.
          - "ja_fechado"  -> o campo 'Estado' ja estava 'Resolvido';
                              o chamado foi pulado sem mexer em nada.
          - "falha"       -> algo deu errado e o chamado nao foi confirmado.
        """
        if not self.abrir_chamado(item.numero):
            return "falha"

        self.aguardar_pagina_completa()

        if not self.entrar_frame_formulario():
            log.warning("Sem acesso ao formulario do chamado -- pulando esse chamado.")
            return "falha"

        self.aguardar_formulario_pronto()

        estado_atual = self.obter_texto_estado()
        if estado_atual == TEXTO_ESTADO_RESOLVIDO:
            log.info(
                "Chamado %s ja esta com Estado = 'Resolvido' -- pulando sem alterar nada.",
                item.numero,
            )
            self.sair_frame_formulario()
            return "ja_fechado"

        if item.grupo:
            self.preencher_autocomplete_classic(ID_CAMPO_GRUPO, item.grupo)
        else:
            log.info("Nenhum valor na coluna 'grupo' pra esse chamado -- pulando esse campo.")

        if item.fechar:
            self.preencher_autocomplete_classic(ID_CAMPO_FECHAR, item.fechar)
        else:
            log.info("Nenhum valor na coluna 'fechar' pra esse chamado -- pulando esse campo.")

        self.aplicar_template()
        resolvido = self.clicar_resolver()

        self.sair_frame_formulario()
        return "fechado" if resolvido else "falha"


# =========================
# ESCRITA DO STATUS DE VOLTA NA PLANILHA
# =========================
PREENCHIMENTO_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PREENCHIMENTO_JA_FECHADO = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
PREENCHIMENTO_FALHA = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def preparar_coluna_status(planilha: Path, sheet: str, header_row: int, nome_coluna: str = COLUNA_STATUS):
    """Abre a planilha com openpyxl (preservando formatacao/formulas) e
    garante que existe uma coluna de status na aba. Se ja existir (mesmo
    nome, ignorando acentos/maiusculas), reaproveita; senao cria uma nova
    coluna logo apos a ultima usada na linha de cabecalho.

    Retorna (workbook, worksheet, indice_da_coluna_status)."""
    try:
        wb = load_workbook(planilha)
    except InvalidFileException as e:
        log.error("Nao consegui abrir a planilha com openpyxl: %s", e)
        raise SystemExit(1)

    if sheet not in wb.sheetnames:
        log.error("A aba '%s' nao existe na planilha.", sheet)
        raise SystemExit(1)

    ws = wb[sheet]
    linha_cabecalho_excel = header_row + 1  # openpyxl e 1-indexado

    ultima_coluna = 0
    col_status = None
    for celula in ws[linha_cabecalho_excel]:
        if celula.value is not None and str(celula.value).strip() != "":
            ultima_coluna = celula.column
            if normalizar(celula.value) == normalizar(nome_coluna):
                col_status = celula.column

    if col_status is None:
        col_status = ultima_coluna + 1
        ws.cell(row=linha_cabecalho_excel, column=col_status, value=nome_coluna)
        log.info("Coluna '%s' criada na planilha (coluna %d).", nome_coluna, col_status)

    return wb, ws, col_status


def atualizar_status_planilha(wb, ws, col_status: int, item: "Chamado", status: str) -> None:
    """Escreve o resultado de um chamado na linha correspondente.

    - status == "fechado": o script resolveu o chamado agora. Grava
      "fechado" e pinta de verde. E esse texto que faz a linha ser
      pulada numa proxima rodada (ver carregar_chamados).
    - status == "ja_fechado": o campo 'Estado' ja estava 'Resolvido' --
      o script nao mexeu em nada. Grava "ja fechado" e pinta de azul.
      Tambem faz a linha ser pulada numa proxima rodada.
    - status == "falha": NAO escreve nada na coluna de status (fica
      vazia), so pinta a celula de vermelho como aviso visual. Assim,
      como a celula continua vazia, o chamado volta a ser tentado
      automaticamente na proxima execucao do script.
    """
    celula = ws.cell(row=item.linha_excel, column=col_status)
    if status == "fechado":
        celula.value = "fechado"
        celula.fill = PREENCHIMENTO_OK
    elif status == "ja_fechado":
        celula.value = "ja fechado"
        celula.fill = PREENCHIMENTO_JA_FECHADO
    else:
        celula.value = None
        celula.fill = PREENCHIMENTO_FALHA


def salvar_planilha_com_retentativa(wb, planilha: Path, tentativas: int = 3) -> bool:
    """Tenta salvar a planilha; se estiver aberta no Excel (PermissionError),
    avisa e da a chance de fechar o arquivo antes de tentar de novo."""
    for tentativa in range(1, tentativas + 1):
        try:
            wb.save(planilha)
            return True
        except PermissionError:
            log.warning(
                "Nao consegui salvar o status na planilha -- ela parece estar aberta no Excel. "
                "(tentativa %d/%d)",
                tentativa,
                tentativas,
            )
            input("  Feche o arquivo no Excel e pressione ENTER pra tentar salvar de novo...")
    log.error("Nao foi possivel salvar o status na planilha apos %d tentativas.", tentativas)
    return False


def criar_driver(caminho_driver: Path) -> webdriver.Edge:
    if not caminho_driver.exists():
        log.error("msedgedriver.exe nao encontrado em: %s", caminho_driver)
        raise SystemExit(1)

    service = Service(str(caminho_driver))
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")

    # --- Reduz os sinais de que o navegador esta sendo controlado por automacao ---
    # Por padrao o Selenium liga uma flag interna do Chromium/Edge
    # (navigator.webdriver = true) e mostra uma barra "Chrome/Edge esta sendo
    # controlado por software de teste automatizado". O Azure AD (login da
    # Microsoft) usa esses sinais pra identificar que e um robo e, por
    # seguranca, pula telas opcionais do fluxo de login -- entre elas a de
    # "Manter conectado?". As linhas abaixo desligam esses sinais.
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    # Nao usamos --user-data-dir/profile-directory de proposito: o
    # ServiceNow autentica pelo login do Windows (SSO), entao o Edge ja
    # abre logado na conta certa sem precisar reaproveitar nenhum perfil
    # salvo. Isso tambem evita o erro "Chrome instance exited" que acontece
    # quando se tenta abrir a pasta "User Data" real com o Edge normal
    # ja rodando.

    driver = webdriver.Edge(service=service, options=options)

    # Camada extra: mesmo com as flags acima, o Edge ainda expoe
    # "navigator.webdriver" como True em algumas versoes. Esse comando roda
    # um pequeno script ANTES de qualquer pagina carregar, sobrescrevendo
    # essa propriedade pra "undefined" -- como se fosse um navegador normal.
    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": (
                    "Object.defineProperty(navigator, 'webdriver', "
                    "{get: () => undefined})"
                )
            },
        )
    except Exception:
        # Se o comando CDP nao for suportado nessa versao do driver, segue
        # sem ele -- as flags do EdgeOptions ja ajudam bastante sozinhas.
        log.warning("Nao foi possivel aplicar o script anti-deteccao via CDP (seguindo sem ele).")

    return driver


# =========================
# PROGRAMA PRINCIPAL
# =========================
def parse_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fecha chamados no ServiceNow a partir de uma planilha.")
    parser.add_argument("--planilha", type=Path, default=PLANILHA_PADRAO, help="Caminho da planilha .xlsx")
    parser.add_argument("--sheet", type=str, default=SHEET_PADRAO, help="Nome da aba a ser lida")
    parser.add_argument("--driver", type=Path, default=DRIVER_PADRAO, help="Caminho do msedgedriver.exe")
    return parser.parse_args()


def main() -> None:
    args = parse_argumentos()

    verificar_planilha_fechada(args.planilha)
    chamados, header_row = carregar_chamados(args.planilha, args.sheet)

    log.info("=" * 50)
    log.info("Total de chamados encontrados na planilha (%s): %d", args.sheet, len(chamados))
    log.info("=" * 50)

    if not chamados:
        log.info("Nenhum chamado para processar. Encerrando.")
        return

    wb, ws, col_status = preparar_coluna_status(args.planilha, args.sheet, header_row)

    resultados: list[tuple[Chamado, str]] = []
    driver = criar_driver(args.driver)
    try:
        automacao = AutomacaoServiceNow(driver)
        for item in chamados:
            status_resultado = automacao.processar_chamado(item)
            resultados.append((item, status_resultado))

            # Grava o status na planilha logo apos cada chamado, pra nao
            # perder o progresso se o script parar no meio (Ctrl+C, erro, etc.)
            atualizar_status_planilha(wb, ws, col_status, item, status_resultado)
            salvar_planilha_com_retentativa(wb, args.planilha)

            # Segue direto pro proximo chamado, sem pausa manual, assim que
            # a pagina volta apos resolver o chamado atual.
    except KeyboardInterrupt:
        log.warning("Interrompido pelo usuario (Ctrl+C).")
    finally:
        # Garante que o navegador fecha mesmo se algo der errado no meio do loop
        driver.quit()
        # Ultima tentativa de salvar, pra garantir que nada de status fique
        # so em memoria se algo tiver dado errado durante o loop.
        salvar_planilha_com_retentativa(wb, args.planilha)

    nao_confirmados = [item.numero for item, status in resultados if status == "falha"]
    ja_estavam_fechados = sum(1 for _, status in resultados if status == "ja_fechado")

    log.info("=" * 50)
    log.info("Concluido! %d/%d chamados processados com confirmacao.", len(resultados) - len(nao_confirmados), len(chamados))
    if ja_estavam_fechados:
        log.info("%d chamado(s) ja estavam com Estado 'Resolvido' e foram pulados.", ja_estavam_fechados)
    if nao_confirmados:
        log.warning("%d chamado(s) NAO confirmados -- verifique manualmente:", len(nao_confirmados))
        for n in nao_confirmados:
            log.warning("  - %s", n)
    log.info("Status gravado direto na planilha (coluna '%s').", COLUNA_STATUS)
    log.info("=" * 50)


if __name__ == "__main__":
    main()